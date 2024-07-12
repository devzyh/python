"""
作者：devzyh
日期：2024-07-12
描述：企业微信在线表格转为邮件内容
"""
import configparser
import json
import os
import time

import requests
from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from selenium import webdriver

# 加载配置文件
config = configparser.ConfigParser()
config.read("config.ini", "UTF-8")

# 常量
common = "common"
tmp_file = "tmp.weekly.xlsx"
item_split = config.get(common, "item_split")
doc_id = config.get(common, "doc_id")


# 解析工作表
def parse_sheet(ws: Worksheet) -> map:
    projects = {}
    for row in ws.iter_rows(values_only=True):
        # 只获取指定用户
        if row[7] != config.get(common, "author"):
            continue

        project_name = row[0]
        project_items = []
        if project_name in projects:
            project_items = projects.get(project_name)
        remarks = row[10]
        if remarks is None:
            remarks = ""
        project_items.append(row[5] + item_split + remarks)

        projects[project_name] = project_items

    return projects


# 输出内容
def print_content(title: str, projects: map):
    print(title + "：")
    for pn in projects:
        print("    " + pn + "：")
        for index, item in enumerate(projects[pn], start=1):
            text = item.split(item_split)
            tb = text[0]
            remarks = text[1]
            if len(remarks) > 0:
                print("        " + str(index) + ". " + tb + "：")
            else:
                print("        " + str(index) + ". " + tb)
                continue
            for r_index, remark in enumerate(remarks.split("\n"), start=1):
                # 目前使用内容自带的序号，不写时替换为变量r_index
                if ". " in remark:
                    print("            " + remark)
                else:
                    print("            " + str(r_index) + ". " + remark)


def download_doc():
    driver = webdriver.Edge()
    driver.get("https://doc.weixin.qq.com/smartsheet/" + doc_id)
    cookies = driver.get_cookies()
    cookie_str = ""
    for cookie in cookies:
        cookie_str = cookie_str + (cookie.get("name") + "=" + cookie.get("value") + "; ")
    print("获取到的Cookies：" + cookie_str)
    driver.quit()

    headers = {"cookie": cookie_str}
    response = requests.post("https://doc.weixin.qq.com/v1/export/export_office?version=2&docId=" + doc_id,
                             headers=headers)
    if response.status_code != 200:
        print("获取请求操作ID失败")
        exit()
    operation_id = json.loads(response.text).get("operationId")
    print("操作ID：" + operation_id)

    progress = 0
    while progress < 100:
        response = requests.get("https://doc.weixin.qq.com/v1/export/query_progress?operationId=" + operation_id,
                                headers=headers)
        if response.status_code != 200:
            print("获取文件下载地址失败")
            exit()

        data = json.loads(response.text)
        file_url = data.get("file_url")
        progress = data.get("progress")
        if progress < 100:
            time.sleep(1)

    print("文件下载地址" + file_url)

    # 导出到临时文件
    response = requests.get(file_url)
    if response.status_code != 200:
        print("导出文件下载失败")
        exit()

    with open(tmp_file, "wb") as f:
        f.write(response.content)
    print("下载文件到临时文件：" + tmp_file)


# 主程序
download_doc()
print("开始提取文件[" + tmp_file + "]内容")

# 读取excel文件
wb = load_workbook(filename=tmp_file)

# 删除临时文件
os.remove(tmp_file)
print("临时文件已删除[" + tmp_file + "]")

print("================================生成内容仅供参考，请根据实际情况进行修改================================")

# 总结
projects = parse_sheet(wb.worksheets[1])
print_content("总结", projects)

print("")

# 计划
projects = parse_sheet(wb.worksheets[0])
print_content("计划", projects)

print("====================================内容已生成完毕，请复制到邮件发送====================================")
input("Press Enter to continue...")